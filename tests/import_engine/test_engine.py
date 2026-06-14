import os
import tempfile
import pytest
from openpyxl import Workbook

from app.import_engine.engine import ImportEngine, ImportResult, SheetImportResult
from app.import_engine.classifier import TypeConfig, ClassificationResult


class MockModel:
    装置名称: str = ""
    位号: str = ""
    设备名称: str = ""
    设备等级: str = ""
    型号规格: str = ""
    生产厂家: str = ""
    安装位置及用途: str = ""
    测量范围: str = ""
    连接方式及规格: str = ""
    精度: str = ""
    出厂编号: str = ""
    是否联锁: str = ""

    __tablename__ = "mock_table"

    def __init__(self, **kwargs):
        for k, v in kwargs.items():
            if hasattr(self, k):
                setattr(self, k, v)


class MockLedger:
    def __init__(self, **kwargs):
        for k, v in kwargs.items():
            setattr(self, k, v)


def _make_simple_excel(filepath, sheet_name, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
    wb.save(filepath)


class TestImportEngine:
    def _make_engine(self):
        engine = ImportEngine()
        engine._types_config = {
            "pressure_remote": {
                "code": "pressure_transmitter",
                "name": "压力变送器",
                "model_class": "MockModel",
                "column_mapping": {
                    "位号": "位号",
                    "设备名称": "设备名称",
                    "设备等级": "设备等级",
                    "安装位置及用途": "安装位置及用途",
                },
            }
        }
        return engine

    def test_import_simple_file(self, tmp_path):
        filepath = os.path.join(tmp_path, "test.xlsx")
        _make_simple_excel(
            filepath, "压力变送器",
            ["序号", "位号", "设备名称", "设备等级"],
            [
                [1, "PT-001", "压力变送器", "A"],
                [2, "PT-002", "压力变送器", "B"],
            ],
        )

        engine = ImportEngine()
        engine._load_configs = lambda: None
        engine._classifier._configs = [
            TypeConfig(
                key="pressure_remote", code="pressure_transmitter",
                name="压力变送器",
                sheet_keywords=["压力变送器", "远传压力"],
                column_signatures=[],
            )
        ]
        engine._types_config = {
            "pressure_remote": {
                "code": "pressure_transmitter",
                "name": "压力变送器",
                "model_class": MockModel,
                "column_mapping": {
                    "位号": "位号",
                    "设备名称": "设备名称",
                    "设备等级": "设备等级",
                },
            }
        }

        result = engine.import_file(filepath)
        assert len(result.sheets) == 1
        sheet_result = result.sheets[0]
        assert sheet_result.sheet_name == "压力变送器"
        assert sheet_result.type_code == "pressure_transmitter"

    def test_import_unknown_type_skipped(self, tmp_path):
        filepath = os.path.join(tmp_path, "unknown.xlsx")
        _make_simple_excel(
            filepath, "未知类型",
            ["序号", "位号"],
            [[1, "TAG-001"]],
        )

        engine = self._make_engine()
        result = engine.import_file(filepath)
        assert len(result.sheets) == 1
        assert result.sheets[0].type_code is None

    def test_multiple_sheets(self, tmp_path):
        filepath = os.path.join(tmp_path, "multi.xlsx")
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "压力变送器"
        for ci, h in enumerate(["序号", "位号"], 1):
            ws1.cell(row=1, column=ci, value=h)
        ws1.cell(row=2, column=1, value=1)
        ws1.cell(row=2, column=2, value="PT-001")

        ws2 = wb.create_sheet("调节阀")
        for ci, h in enumerate(["序号", "位号"], 1):
            ws2.cell(row=1, column=ci, value=h)
        ws2.cell(row=2, column=1, value=1)
        ws2.cell(row=2, column=2, value="CV-001")
        wb.save(filepath)

        engine = ImportEngine()
        engine._load_configs = lambda: None
        engine._classifier._configs = [
            TypeConfig(
                key="pressure_remote", code="pressure_transmitter",
                name="压力变送器",
                sheet_keywords=["压力变送器"],
                column_signatures=[],
            ),
            TypeConfig(
                key="control_valve", code="valve",
                name="阀门",
                sheet_keywords=["调节阀"],
                column_signatures=[],
            ),
        ]
        engine._types_config = {
            "pressure_remote": {
                "code": "pressure_transmitter",
                "name": "压力变送器",
                "model_class": MockModel,
                "column_mapping": {"位号": "位号"},
            },
            "control_valve": {
                "code": "valve",
                "name": "阀门",
                "model_class": MockModel,
                "column_mapping": {"位号": "位号"},
            },
        }

        result = engine.import_file(filepath)
        assert len(result.sheets) == 2

    def test_summary_sheet_skipped_from_classification(self, tmp_path):
        filepath = os.path.join(tmp_path, "with_summary.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "汇总"
        ws.cell(row=1, column=1, value="装置名称")
        ws.cell(row=1, column=2, value="A级")
        wb.save(filepath)
        ws2 = wb.create_sheet("压力变送器")
        ws2.cell(row=1, column=1, value="序号")
        ws2.cell(row=1, column=2, value="位号")
        ws2.cell(row=2, column=1, value=1)
        ws2.cell(row=2, column=2, value="PT-001")
        wb.save(filepath)

        engine = ImportEngine()
        engine._load_configs = lambda: None
        engine._classifier._configs = [
            TypeConfig(
                key="pressure_remote", code="pressure_transmitter",
                name="压力变送器",
                sheet_keywords=["压力变送器"],
                column_signatures=[],
            ),
            TypeConfig(
                key="summary", code="summary",
                name="汇总表",
                sheet_keywords=["汇总"],
                column_signatures=[],
            ),
        ]
        engine._types_config = {
            "pressure_remote": {
                "code": "pressure_transmitter",
                "name": "压力变送器",
                "model_class": MockModel,
                "column_mapping": {"位号": "位号"},
            },
        }

        result = engine.import_file(filepath)
        pressure_sheets = [s for s in result.sheets if s.type_code == "pressure_transmitter"]
        assert len(pressure_sheets) == 1
        summary_sheets = [s for s in result.sheets if s.type_code == "summary"]
        assert len(summary_sheets) == 1

    def test_verification_with_summary_sheet(self, tmp_path):
        filepath = os.path.join(tmp_path, "verify.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "汇总"
        ws.cell(row=1, column=1, value="装置名称")
        ws.cell(row=1, column=2, value="A级")
        ws.cell(row=1, column=3, value="B级")
        ws.cell(row=1, column=4, value="C级")
        ws.cell(row=2, column=1, value="装置1")
        ws.cell(row=2, column=2, value=5)
        ws.cell(row=2, column=3, value=3)
        ws.cell(row=2, column=4, value=2)

        ws2 = wb.create_sheet("压力变送器")
        ws2.cell(row=1, column=1, value="序号")
        ws2.cell(row=1, column=2, value="位号")
        ws2.cell(row=1, column=3, value="设备等级")
        ws2.cell(row=2, column=1, value=1)
        ws2.cell(row=2, column=2, value="PT-001")
        ws2.cell(row=2, column=3, value="A")
        ws2.cell(row=3, column=1, value=2)
        ws2.cell(row=3, column=2, value="PT-002")
        ws2.cell(row=3, column=3, value="B")
        wb.save(filepath)

        engine = ImportEngine()
        engine._load_configs = lambda: None
        engine._classifier._configs = [
            TypeConfig(
                key="pressure_remote", code="pressure_transmitter",
                name="压力变送器",
                sheet_keywords=["压力变送器"],
                column_signatures=[],
            ),
            TypeConfig(
                key="summary", code="summary",
                name="汇总表",
                sheet_keywords=["汇总"],
                column_signatures=[],
            ),
        ]
        engine._types_config = {
            "pressure_remote": {
                "code": "pressure_transmitter",
                "name": "压力变送器",
                "model_class": MockModel,
                "column_mapping": {"位号": "位号", "设备等级": "设备等级"},
            },
        }

        result = engine.import_file(filepath)
        assert result.verification is not None
        assert result.verification.mismatches is not None
        assert result.summary_data is not None
        assert result.summary_data.sheet_name == "汇总"
