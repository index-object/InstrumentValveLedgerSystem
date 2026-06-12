import os
import tempfile
from unittest.mock import Mock, patch
import pytest
from openpyxl import Workbook

from app.import_engine.extractor import DataExtractor, SheetData, HeaderInfo


def _make_workbook(sheet_name, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    return wb


def _make_multi_header_workbook(sheet_name, top_headers, sub_headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for col_idx, h in enumerate(top_headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    for col_idx, h in enumerate(sub_headers, 1):
        ws.cell(row=2, column=col_idx, value=h)
    for row_idx, row_data in enumerate(rows, 3):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    return wb


class TestDataExtractor:
    def test_detect_single_header_row(self):
        headers = ["序号", "位号", "设备名称", "设备等级"]
        rows = [
            [1, "TAG-001", "压力表", "A"],
            [2, "TAG-002", "温度计", "B"],
        ]
        wb = _make_workbook("Sheet1", headers, rows)
        extractor = DataExtractor()
        result = extractor.extract_sheet(wb["Sheet1"])

        assert result.header_info.header_row == 1
        assert result.header_info.data_start_row == 2
        assert result.headers == headers
        assert len(result.rows) == 2

    def test_detect_double_header_row(self):
        top = ["序号", "工艺条件", "", "", "阀体参数", "", "备注"]
        sub = ["", "介质名称", "设计温度", "阀前压力", "公称通径", "材质", ""]
        rows = [
            [1, "蒸汽", "450", "4.0", "DN100", "WCB", ""],
        ]
        wb = _make_multi_header_workbook("Sheet1", top, sub, rows)
        extractor = DataExtractor()
        result = extractor.extract_sheet(wb["Sheet1"])

        assert result.header_info.is_double_row
        assert result.header_info.data_start_row == 3
        assert "工艺条件.介质名称" in result.headers
        assert "设计温度" in result.headers
        assert len(result.rows) == 1

    def test_filter_accessory_rows(self):
        headers = ["序号", "位号", "设备名称", "设备等级"]
        rows = [
            [1, "TAG-001", "调节阀", "A"],
            [None, "", "执行机构", ""],
            [None, "", "定位器", ""],
            [2, "TAG-002", "开关阀", "B"],
        ]
        wb = _make_workbook("Sheet1", headers, rows)
        extractor = DataExtractor()
        result = extractor.extract_sheet(wb["Sheet1"])

        assert len(result.rows) == 2
        assert len(result.accessories) == 2
        assert result.rows[0]["设备名称"] == "调节阀"
        assert result.rows[1]["设备名称"] == "开关阀"

    def test_sequence_number_float_to_int(self):
        headers = ["序号", "位号"]
        rows = [
            [1.0, "TAG-001"],
            [2.0, "TAG-002"],
            [3, "TAG-003"],
        ]
        wb = _make_workbook("Sheet1", headers, rows)
        extractor = DataExtractor()
        result = extractor.extract_sheet(wb["Sheet1"])

        assert result.rows[0]["序号"] == "1"
        assert result.rows[1]["序号"] == "2"
        assert result.rows[2]["序号"] == "3"

    def test_device_grade_normalization(self):
        headers = ["序号", "位号", "设备等级"]
        rows = [
            [1, "TAG-001", "A"],
            [2, "TAG-002", " B "],
            [3, "TAG-003", "c类"],
            [4, "TAG-004", "C"],
        ]
        wb = _make_workbook("Sheet1", headers, rows)
        extractor = DataExtractor()
        result = extractor.extract_sheet(wb["Sheet1"])

        assert result.rows[1]["设备等级"] == "B"
        assert result.rows[2]["设备等级"] == "C"

    def test_empty_cells_use_empty_string(self):
        headers = ["序号", "位号", "生产厂家"]
        rows = [
            [1, "TAG-001", None],
            [2, "TAG-002", ""],
        ]
        wb = _make_workbook("Sheet1", headers, rows)
        extractor = DataExtractor()
        result = extractor.extract_sheet(wb["Sheet1"])

        assert result.rows[0]["生产厂家"] == ""
        assert result.rows[1]["生产厂家"] == ""

    def test_skip_empty_rows(self):
        headers = ["序号", "位号"]
        rows = [
            [1, "TAG-001"],
            [None, None],
            [None, None],
            [2, "TAG-002"],
        ]
        wb = _make_workbook("Sheet1", headers, rows)
        extractor = DataExtractor()
        result = extractor.extract_sheet(wb["Sheet1"])

        assert len(result.rows) == 2

    def test_summary_sheet_no_filtering(self):
        headers = ["装置名称", "A级", "B级", "C级"]
        rows = [
            ["气化装置", 5, 10, 3],
            ["合成装置", 2, 8, 1],
        ]
        wb = _make_workbook("汇总", headers, rows)
        extractor = DataExtractor()
        result = extractor.extract_sheet(wb["汇总"], is_summary=True)

        assert len(result.rows) == 2
        assert len(result.accessories) == 0

    def test_header_not_found_fallback(self):
        headers = ["列1", "列2", "列3"]
        rows = [["a", "b", "c"]]
        wb = _make_workbook("Sheet1", headers, rows)
        extractor = DataExtractor()
        result = extractor.extract_sheet(wb["Sheet1"])

        assert result.header_info.header_row == 1
        assert len(result.headers) == 3

    def test_skip_completely_empty_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Empty"
        extractor = DataExtractor()
        result = extractor.extract_sheet(ws)

        assert len(result.headers) == 0
        assert len(result.rows) == 0

    def test_read_excel_file(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws.cell(row=1, column=1, value="序号")
        ws.cell(row=1, column=2, value="位号")
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=2, value="TAG-001")
        filepath = os.path.join(tmp_path, "test.xlsx")
        wb.save(filepath)

        extractor = DataExtractor()
        sheets = extractor.read_excel(filepath)
        assert len(sheets) == 1
        assert sheets[0].sheet_name == "TestSheet"
        assert len(sheets[0].rows) == 1
        assert sheets[0].rows[0]["位号"] == "TAG-001"
