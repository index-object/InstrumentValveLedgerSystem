import pytest
import io
from openpyxl import Workbook


class TestParseExcel:
    def test_parse_valid_file(self, tmp_path):
        from app.routes.maintenance_import import _parse_xlsx

        wb = Workbook()
        ws = wb.active
        ws.title = "维护记录"
        ws.append(["装置名称", "设备位号", "设备名称", "检修时间", "检修内容", "检修人员", "类型"])
        ws.append(["气化装置", "PT-001", "压力变送器", "2025-03-15", "更换膜片", "张三", "大修"])
        ws.append(["气化装置", "PT-002", "压力变送器", "2025-04-01", "校准", "李四", "小修"])
        filepath = str(tmp_path / "test.xlsx")
        wb.save(filepath)
        wb.close()

        records = _parse_xlsx(filepath)
        assert len(records) == 2
        assert records[0]["装置名称"] == "气化装置"
        assert records[0]["设备位号"] == "PT-001"
        assert records[0]["检修人员"] == "张三"

    def test_parse_empty_rows_skipped(self, tmp_path):
        from app.routes.maintenance_import import _parse_xlsx

        wb = Workbook()
        ws = wb.active
        ws.append(["装置名称", "设备位号"])
        ws.append(["气化装置", "PT-001"])
        ws.append([None, None])
        ws.append(["", ""])
        ws.append(["合成装置", "TT-002"])
        filepath = str(tmp_path / "test.xlsx")
        wb.save(filepath)
        wb.close()

        records = _parse_xlsx(filepath)
        assert len(records) == 2

    def test_parse_missing_required_columns(self, tmp_path):
        from app.routes.maintenance_import import _parse_xlsx

        wb = Workbook()
        ws = wb.active
        ws.append(["装置名称", "设备位号"])
        ws.append(["", "PT-001"])
        ws.append(["气化装置", ""])
        ws.append(["气化装置", "PT-002"])
        filepath = str(tmp_path / "test.xlsx")
        wb.save(filepath)
        wb.close()

        records = _parse_xlsx(filepath)
        assert len(records) == 1
        assert records[0]["设备位号"] == "PT-002"

    def test_parse_header_only_file(self, tmp_path):
        from app.routes.maintenance_import import _parse_xlsx

        wb = Workbook()
        ws = wb.active
        ws.append(["装置名称", "设备位号"])
        filepath = str(tmp_path / "test.xlsx")
        wb.save(filepath)
        wb.close()

        records = _parse_xlsx(filepath)
        assert len(records) == 0

    def test_parse_empty_file(self, tmp_path):
        from app.routes.maintenance_import import _parse_xlsx

        wb = Workbook()
        ws = wb.active
        filepath = str(tmp_path / "test.xlsx")
        wb.save(filepath)
        wb.close()

        records = _parse_xlsx(filepath)
        assert len(records) == 0


class TestImportIntegration:
    """纯单元测试，绕过 fixtures 和数据库，直接测试 resolve_device 和 import 的逻辑函数"""

    def test_resolve_device_missing_args(self):
        from app.utils.device_lookup import resolve_device
        assert resolve_device("装置", "/") is None
        assert resolve_device("装置", "") is None
        assert resolve_device("装置", "-") is None
        assert resolve_device("装置", "\\") is None
        assert resolve_device("", "TAG") is None
        assert resolve_device(None, "TAG") is None
        assert resolve_device("装置", None) is None

    def test_parse_datetime(self):
        from app.routes.maintenance_import import _parse_datetime
        from datetime import datetime

        d = _parse_datetime("2025-03-15")
        assert d is not None
        assert d.year == 2025 and d.month == 3 and d.day == 15

        d = _parse_datetime("2025-03-15 14:30:00")
        assert d is not None
        assert d.hour == 14 and d.minute == 30

        d = _parse_datetime("")
        assert d is None

        d = _parse_datetime(None)
        assert d is None

        d = _parse_datetime("2025/03/15")
        assert d is not None
        assert d.year == 2025 and d.month == 3 and d.day == 15

    def test_get_type_name(self):
        from app.routes.maintenance_import import _get_type_name
        name = _get_type_name("control_valve")
        assert name is not None
        assert isinstance(name, str)
        assert len(name) > 0

    def test_get_type_name_unknown(self):
        from app.routes.maintenance_import import _get_type_name
        assert _get_type_name("nonexistent_type") == "nonexistent_type"

    def test_check_duplicate_none_when_no_dt(self):
        from app.routes.maintenance_import import _check_duplicate
        assert _check_duplicate("装置", "TAG", "") is False
        assert _check_duplicate("装置", "TAG", None) is False

    def test_duplicate_filtered_in_preview(self, app):
        from app.routes.maintenance_import import _check_duplicate
        with app.app_context():
            assert _check_duplicate("x", "y", "2025-01-01") is False
