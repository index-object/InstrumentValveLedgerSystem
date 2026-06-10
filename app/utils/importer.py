import zipfile
import tempfile
import os
import re
from openpyxl import load_workbook


def safe_read_excel(filepath):
    """安全读取 Excel，移除 externalLinks 避免 openpyxl 挂起"""
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    try:
        with zipfile.ZipFile(filepath) as zin:
            with zipfile.ZipFile(tmp.name, "w") as zout:
                for item in zin.infolist():
                    if item.filename.startswith("xl/externalLinks"):
                        continue
                    if item.filename == "xl/workbook.xml":
                        content = zin.read(item.filename).decode("utf-8")
                        content = re.sub(
                            r'<externalReferences[^>]*>.*?</externalReferences>',
                            "",
                            content,
                            flags=re.DOTALL,
                        )
                        zout.writestr(item, content.encode("utf-8"))
                    else:
                        zout.writestr(item, zin.read(item.filename))
        wb = load_workbook(tmp.name, read_only=True, data_only=True)
        result = []
        for s in wb.sheetnames:
            ws = wb[s]
            headers = []
            rows_data = []
            row_count = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c).strip() if c else "" for c in row]
                else:
                    row_dict = {}
                    for j, cell in enumerate(row):
                        col_name = headers[j] if j < len(headers) else f"col{j}"
                        row_dict[col_name] = str(cell).strip() if cell is not None else ""
                    if any(v for v in row_dict.values()):
                        rows_data.append(row_dict)
                        row_count += 1
            result.append({
                "sheet": s,
                "columns": headers,
                "rows": rows_data,
                "row_count": row_count,
                "sample": rows_data[:5],
            })
        wb.close()
        return result
    finally:
        try:
            os.unlink(tmp.name)
        except:
            pass
