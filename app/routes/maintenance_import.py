import os
import uuid
from datetime import datetime

from flask import (
    Blueprint, render_template, request, flash,
    redirect, url_for, session, current_app,
)
from flask_login import login_required, current_user
from openpyxl import load_workbook

from app.models import db, MaintenanceRecord, Setting
from app.utils.device_lookup import resolve_device
from app.devices import DeviceTypeRegistry
from app.utils.import_cache import cleanup_import_cache

maintenance_import = Blueprint("maintenance_import", __name__,
                               template_folder="../../templates")


@maintenance_import.route("/maintenance/import")
@login_required
def index():
    skipped = session.pop("maintenance_import_skipped", None)
    return render_template("maintenance_import/import.html", skipped=skipped)


def _parse_xlsx(filepath: str) -> list[dict]:
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        wb.close()
        return []

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows or len(rows) < 2:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    records = []
    for row_vals in rows[1:]:
        if not row_vals or all(v is None or str(v).strip() == "" for v in row_vals):
            continue
        record = {}
        for i, val in enumerate(row_vals):
            if i < len(headers) and headers[i]:
                record[headers[i]] = str(val).strip() if val is not None else ""
        if record.get("装置名称") and record.get("设备位号"):
            records.append(record)
    return records


@maintenance_import.route("/maintenance/import/upload", methods=["POST"])
@login_required
def upload():
    if "file" not in request.files:
        flash("请选择文件")
        return redirect(url_for("maintenance_import.index"))

    file = request.files["file"]
    if file.filename == "":
        flash("请选择文件")
        return redirect(url_for("maintenance_import.index"))

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in (".xlsx", ".xls"):
        flash("仅支持 .xlsx / .xls 文件")
        return redirect(url_for("maintenance_import.index"))

    uid = uuid.uuid4().hex
    saved_name = f"maintenance_import_{uid}{ext}"
    upload_folder = current_app.config.get("UPLOAD_FOLDER")
    saved_path = os.path.join(upload_folder, saved_name)
    file.save(saved_path)

    retention = Setting.query.get("import_cache_retention")
    max_keep = int(retention.value) if retention else 30
    cleanup_import_cache(upload_folder, max_keep)

    try:
        raw_records = _parse_xlsx(saved_path)
    except Exception as e:
        flash(f"文件解析失败: {e}")
        try:
            os.remove(saved_path)
        except Exception:
            pass
        return redirect(url_for("maintenance_import.index"))

    if not raw_records:
        flash("文件中没有有效数据（至少需要装置名称和设备位号）")
        try:
            os.remove(saved_path)
        except Exception:
            pass
        return redirect(url_for("maintenance_import.index"))

    matched = []
    unmatched = []
    ambiguous = []
    duplicates = []

    for idx, rec in enumerate(raw_records):
        unit = rec.get("装置名称", "")
        tag = rec.get("设备位号", "")
        if _check_duplicate(unit, tag, rec.get("检修时间", "")):
            duplicates.append({"index": idx, "data": rec})
            continue
        result = resolve_device(unit, tag)
        if result is None:
            unmatched.append({"index": idx, "data": rec})
        elif len(result) == 1:
            dev_type, dev_id, snapshot = result[0]
            matched.append({
                "index": idx,
                "data": rec,
                "device_type": dev_type,
                "device_id": dev_id,
                "snapshot": snapshot,
            })
        else:
            ambiguous.append({
                "index": idx,
                "data": rec,
                "candidates": [
                    {"code": code, "name": _get_type_name(code)}
                    for code, _, _ in result
                ],
            })

    session["maintenance_import_file"] = saved_name
    session["maintenance_import_filename"] = file.filename
    session["maintenance_import_raw"] = raw_records
    session["maintenance_import_matched"] = matched
    session["maintenance_import_unmatched"] = unmatched
    session["maintenance_import_ambiguous"] = ambiguous
    session["maintenance_import_duplicates"] = duplicates

    return render_template(
        "maintenance_import/preview.html",
        filename=file.filename,
        total=len(raw_records),
        matched=matched,
        unmatched=unmatched,
        ambiguous=ambiguous,
        duplicates=duplicates,
        matched_count=len(matched),
        unmatched_count=len(unmatched),
        ambiguous_count=len(ambiguous),
        duplicate_count=len(duplicates),
        has_ambiguous=len(ambiguous) > 0,
        has_unmatched=len(unmatched) > 0,
        has_duplicates=len(duplicates) > 0,
    )


@maintenance_import.route("/maintenance/import/execute", methods=["POST"])
@login_required
def execute():
    saved_name = session.get("maintenance_import_file")
    if not saved_name:
        flash("找不到已上传的文件，请重新上传")
        return redirect(url_for("maintenance_import.index"))

    upload_folder = current_app.config.get("UPLOAD_FOLDER")
    saved_path = os.path.join(upload_folder, saved_name)

    raw_records = session.get("maintenance_import_raw", [])
    matched = session.get("maintenance_import_matched", [])
    unmatched = session.get("maintenance_import_unmatched", [])
    ambiguous = session.get("maintenance_import_ambiguous", [])
    duplicates = session.get("maintenance_import_duplicates", [])

    unmatched_action = request.form.get("unmatched_action", "skip")

    amb_choices = {}
    for key, value in request.form.items():
        if key.startswith("ambiguous_choice_"):
            idx = int(key[len("ambiguous_choice_"):])
            amb_choices[idx] = value

    created = 0
    skipped = 0
    skipped_details = []

    # ponytail: duplicates already excluded from matched/unmatched/ambiguous
    for dup in duplicates:
        skipped += 1
        skipped_details.append({
            "设备位号": dup["data"].get("设备位号", ""),
            "装置名称": dup["data"].get("装置名称", ""),
            "原因": "与数据库已有记录重复（装置+位号+检修日期一致）",
        })

    for m in matched:
        idx = m["index"]
        rec = raw_records[idx]
        dev_type = m["device_type"]
        dev_id = m["device_id"]
        snapshot = m["snapshot"]

        record = MaintenanceRecord(
            device_type=dev_type,
            device_id=dev_id,
            装置名称=snapshot["装置名称"],
            设备位号=snapshot["设备位号"],
            设备名称=snapshot["设备名称"],
            检修时间=_parse_datetime(rec.get("检修时间", "")),
            检修内容=rec.get("检修内容", ""),
            检修人员=rec.get("检修人员", ""),
            类型=rec.get("类型", ""),
            created_by=current_user.id,
        )
        db.session.add(record)
        created += 1

    for amb in ambiguous:
        amb_idx = amb["index"]
        choice = amb_choices.get(amb_idx)
        if not choice:
            skipped += 1
            skipped_details.append({
                "设备位号": amb["data"].get("设备位号", ""),
                "装置名称": amb["data"].get("装置名称", ""),
                "原因": "歧义行未选择类型，已跳过",
            })
            continue
        rec = raw_records[amb_idx]
        resolved = resolve_device(rec.get("装置名称", ""), rec.get("设备位号", ""))
        if not resolved:
            skipped += 1
            skipped_details.append({
                "设备位号": rec.get("设备位号", ""),
                "装置名称": rec.get("装置名称", ""),
                "原因": "歧义行解析失败，已跳过",
            })
            continue
        match = None
        for code, did, snap in resolved:
            if code == choice:
                match = (code, did, snap)
                break
        if not match:
            skipped += 1
            skipped_details.append({
                "设备位号": rec.get("设备位号", ""),
                "装置名称": rec.get("装置名称", ""),
                "原因": "歧义行选择类型后匹配失败，已跳过",
            })
            continue
        dev_type, dev_id, snapshot = match
        record = MaintenanceRecord(
            device_type=dev_type,
            device_id=dev_id,
            装置名称=snapshot["装置名称"],
            设备位号=snapshot["设备位号"],
            设备名称=snapshot["设备名称"],
            检修时间=_parse_datetime(rec.get("检修时间", "")),
            检修内容=rec.get("检修内容", ""),
            检修人员=rec.get("检修人员", ""),
            类型=rec.get("类型", ""),
            created_by=current_user.id,
        )
        db.session.add(record)
        created += 1

    if unmatched_action != "skip":
        for um in unmatched:
            idx = um["index"]
            rec = raw_records[idx]
            record = MaintenanceRecord(
                device_type="",
                device_id=0,
                valve_deleted=True,
                装置名称=rec.get("装置名称", ""),
                设备位号=rec.get("设备位号", ""),
                设备名称=rec.get("设备名称", ""),
                检修时间=_parse_datetime(rec.get("检修时间", "")),
                检修内容=rec.get("检修内容", ""),
                检修人员=rec.get("检修人员", ""),
                类型=rec.get("类型", ""),
                created_by=current_user.id,
            )
            db.session.add(record)
            created += 1
    else:
        for um in unmatched:
            skipped += 1
            skipped_details.append({
                "设备位号": um["data"].get("设备位号", ""),
                "装置名称": um["data"].get("装置名称", ""),
                "原因": "未找到匹配设备，已跳过",
            })

    db.session.commit()

    try:
        os.remove(saved_path)
    except Exception:
        pass
    for key in (
        "maintenance_import_file", "maintenance_import_filename",
        "maintenance_import_raw", "maintenance_import_matched",
        "maintenance_import_unmatched", "maintenance_import_ambiguous",
        "maintenance_import_duplicates",
    ):
        session.pop(key, None)

    parts = [f"创建 {created} 条"]
    if skipped:
        parts.append(f"跳过 {skipped} 条")
    flash(f"导入完成：{'，'.join(parts)}")
    if skipped_details:
        session["maintenance_import_skipped"] = skipped_details
    return redirect(url_for("maintenance_import.index"))


def _check_duplicate(unit: str, tag: str, time_str: str) -> bool:
    dt = _parse_datetime(time_str)
    if not dt:
        return False
    return MaintenanceRecord.query.filter(
        MaintenanceRecord.装置名称 == unit,
        MaintenanceRecord.设备位号 == tag,
        MaintenanceRecord.检修时间 == dt,
    ).first() is not None


def _get_type_name(code: str) -> str:
    config = DeviceTypeRegistry.get(code)
    return config.name if config else code


def _parse_datetime(s: str):
    if not s or not s.strip():
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
                "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y/%m/%d",
                "%Y.%m.%d"):
        try:
            return datetime.strptime(s.strip(), fmt)
        except ValueError:
            continue
    return None
