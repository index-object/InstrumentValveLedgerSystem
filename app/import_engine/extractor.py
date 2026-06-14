from dataclasses import dataclass, field
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class _XlrdCell:
    """适配 xlrd cell 为 openpyxl cell 接口"""

    def __init__(self, value: Any, column: int):
        self.value = value
        self.column = column


class _XlrdSheet:
    """将 xlrd sheet 适配为 openpyxl Worksheet 接口"""

    def __init__(self, sheet):
        self._sheet = sheet
        self.title = sheet.name
        self.max_row = sheet.nrows
        self.max_column = sheet.ncols

    def __getitem__(self, row_idx: int):
        """返回 1-based row 的 cell 列表"""
        if row_idx < 1 or row_idx > self.max_row:
            return ()
        row = self._sheet.row(row_idx - 1)
        return tuple(_XlrdCell(cell.value, cell_col + 1) for cell_col, cell in enumerate(row))

    def iter_rows(self, min_row=1, max_row=None, values_only=False):
        """模拟 openpyxl 的 iter_rows"""
        if max_row is None:
            max_row = self.max_row
        start = max(min_row, 1) - 1
        end = min(max_row, self.max_row)
        for row_idx in range(start, end):
            row = self._sheet.row(row_idx)
            if values_only:
                yield tuple(cell.value for cell in row)
            else:
                yield tuple(_XlrdCell(cell.value, cell_col + 1) for cell_col, cell in enumerate(row))


@dataclass
class HeaderInfo:
    header_row: int = 0
    data_start_row: int = 2
    is_double_row: bool = False


@dataclass
class SheetData:
    sheet_name: str
    headers: list[str]
    rows: list[dict[str, str]]
    accessories: list[dict[str, str]] = field(default_factory=list)
    header_info: Optional[HeaderInfo] = None


class DataExtractor:
    SEQUENCE_KEYWORDS = ["序号", "No", "no", "NO", "序列号"]

    def read_excel(self, filepath: str) -> list[SheetData]:
        ext = filepath.rsplit(".", 1)[-1].lower() if "." in filepath else ""
        if ext == "xls":
            return self._read_xls(filepath)
        return self._read_xlsx(filepath)

    def _read_xlsx(self, filepath: str) -> list[SheetData]:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        results = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result = self.extract_sheet(ws)
            results.append(result)
        wb.close()
        return results

    def _read_xls(self, filepath: str) -> list[SheetData]:
        import xlrd
        wb = xlrd.open_workbook(filepath)
        results = []
        for sheet_name in wb.sheet_names():
            xl_sheet = wb.sheet_by_name(sheet_name)
            ws = _XlrdSheet(xl_sheet)
            result = self.extract_sheet(ws)
            results.append(result)
        return results

    def extract_sheet(
        self, ws: Worksheet, is_summary: bool = False
    ) -> SheetData:
        sheet_name = ws.title
        max_row = ws.max_row or 0
        max_col = self._get_max_col(ws)

        if max_row == 0 or max_col == 0:
            return SheetData(sheet_name=sheet_name, headers=[], rows=[])

        header_info = self._detect_headers(ws)
        raw_headers = self._build_headers(ws, header_info, max_col)

        rows, accessories = self._extract_rows(
            ws, header_info, max_row, max_col, raw_headers, is_summary
        )

        return SheetData(
            sheet_name=sheet_name,
            headers=raw_headers,
            rows=rows,
            accessories=accessories,
            header_info=header_info,
        )

    def _detect_headers(self, ws: Worksheet) -> HeaderInfo:
        info = HeaderInfo()
        max_search = min(11, (ws.max_row or 0) + 1)
        for row_idx in range(1, max_search):
            row_values = [
                str(c.value).strip() if c.value is not None else ""
                for c in ws[row_idx]
            ]
            combined = " ".join(row_values)

            if any(kw in combined for kw in self.SEQUENCE_KEYWORDS):
                info.header_row = row_idx
                total_cols = max(len(row_values), 1)
                non_empty = sum(1 for v in row_values if v)
                if non_empty < total_cols * 0.6:
                    info.is_double_row = True
                    info.data_start_row = row_idx + 2
                else:
                    info.data_start_row = row_idx + 1
                return info

        info.header_row = 1
        info.data_start_row = 2
        return info

    def _get_max_col(self, ws: Worksheet) -> int:
        max_col = 0
        for row in ws.iter_rows(min_row=1, max_row=min(11, ws.max_row or 1)):
            for cell in row:
                if cell.value is not None:
                    max_col = max(max_col, cell.column)
        return max_col

    def _build_headers(
        self, ws: Worksheet, header_info: HeaderInfo, max_col: int
    ) -> list[str]:
        if header_info.is_double_row:
            top_row = list(ws[header_info.header_row])
            sub_row_idx = header_info.header_row + 1
            sub_row = list(ws[sub_row_idx]) if sub_row_idx <= (ws.max_row or 0) else []
            headers = []
            for col_idx in range(1, max_col + 1):
                top_val = (
                    str(top_row[col_idx - 1].value).strip()
                    if col_idx <= len(top_row) and top_row[col_idx - 1].value is not None
                    else ""
                )
                sub_val = (
                    str(sub_row[col_idx - 1].value).strip()
                    if col_idx <= len(sub_row) and sub_row[col_idx - 1].value is not None
                    else ""
                )
                if sub_val:
                    if top_val:
                        headers.append(f"{top_val}.{sub_val}")
                    else:
                        headers.append(sub_val)
                else:
                    headers.append(top_val)
            return headers
        else:
            row = list(ws[header_info.header_row])
            return [
                str(c.value).strip() if c.value is not None else ""
                for c in row
            ]

    def _extract_rows(
        self,
        ws: Worksheet,
        header_info: HeaderInfo,
        max_row: int,
        max_col: int,
        headers: list[str],
        is_summary: bool,
    ) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
        rows = []
        accessories = []

        for row_values in ws.iter_rows(
            min_row=header_info.data_start_row,
            max_row=max_row,
            values_only=True,
        ):
            vals = list(row_values)
            if len(vals) < max_col:
                vals.extend([""] * (max_col - len(vals)))
            row_data = [self._clean_cell_value(v) for v in vals[:max_col]]

            if all(v == "" for v in row_data):
                continue

            row_dict = {}
            for col_idx, val in enumerate(row_data):
                col_name = (
                    headers[col_idx]
                    if col_idx < len(headers)
                    else f"col{col_idx}"
                )
                row_dict[col_name] = val

            seq_val = row_data[0] if row_data else ""

            if is_summary:
                rows.append(row_dict)
            elif seq_val:
                cleaned = self._clean_row(row_dict)
                rows.append(cleaned)
            else:
                has_content = any(
                    v.strip() for k, v in row_dict.items() if k and v
                )
                if has_content:
                    accessories.append(row_dict)

        return rows, accessories

    def _clean_cell_value(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float):
            if value == int(value):
                return str(int(value))
        return str(value).strip()

    def _clean_row(self, row: dict[str, str]) -> dict[str, str]:
        cleaned = {}
        for key, val in row.items():
            val = val.strip()
            if self._is_device_grade_field(key):
                val = self._normalize_device_grade(val)
            if self._is_sequence_field(key):
                val = self._normalize_sequence(val)
            cleaned[key] = val
        return cleaned

    def _is_device_grade_field(self, field_name: str) -> bool:
        return "设备等级" in field_name or "设备分级" in field_name or "分级" in field_name

    def _is_sequence_field(self, field_name: str) -> bool:
        return any(kw in field_name for kw in self.SEQUENCE_KEYWORDS)

    def _normalize_device_grade(self, val: str) -> str:
        val = val.replace("类", "").strip()
        val = val.strip().upper()
        if val in ("A", "B", "C"):
            return val
        return val

    def _normalize_sequence(self, val: str) -> str:
        try:
            if "." in val:
                f = float(val)
                if f == int(f):
                    return str(int(f))
            return val
        except ValueError:
            return val
